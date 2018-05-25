# -*- coding: utf-8 -*-

import telebot
import datetime
from get_base_of_Date import *
from keyboard import *
import misk
from datetime import datetime
import cherrypy
from telebot import types
from copy import deepcopy
#from function import *
#from openpyxl import Workbook
from openpyxl import load_workbook
from json import load, dump
token = misk.token
bot = telebot.TeleBot(token)
import requests

# bot.remove_webhook()


# from telebot import apihelper

# apihelper.proxy = {'http', 'http://5.101.64.68:57624'}

# WEBHOOK_HOST = '176.124.146.219'
# WEBHOOK_PORT = 443  # 443, 80, 88 или 8443 (порт должен быть открыт!)
# WEBHOOK_LISTEN = '176.124.146.219'  # На некоторых серверах придется указывать такой же IP, что и выше

# WEBHOOK_SSL_CERT = 'ssl/webhook_cert.pem'  # Путь к сертификату
# WEBHOOK_SSL_PRIV = 'ssl/webhook_pkey.pem'  # Путь к приватному ключу

# WEBHOOK_URL_BASE = "https://%s:%s" % (WEBHOOK_HOST, WEBHOOK_PORT)
# WEBHOOK_URL_PATH = "/%s/" % (misk.token)

class WebhookServer(object):
    @cherrypy.expose
    def index(self):
        if 'content-length' in cherrypy.request.headers and \
                        'content-type' in cherrypy.request.headers and \
                        cherrypy.request.headers['content-type'] == 'application/json':
            length = int(cherrypy.request.headers['content-length'])
            json_string = cherrypy.request.body.read(length).decode("utf-8")
            update = telebot.types.Update.de_json(json_string)
            # Эта функция обеспечивает проверку входящего сообщения
            bot.process_new_updates([update])
            return ''
        else:
            raise cherrypy.HTTPError(403)

month_table = {'1': "Январь",
		'2': "Февраль",
		'3': "Март",
		'4': "Апрель",
		'5': "Май",
		'6': "Июнь",
		'7': "Июль",
		'8': "Август",
		'9': "Сентябрь",
		'10': "Октябрь",
		'11': "Ноябрь",
		'12': "Декабрь"}

weekday = { '1': "ПН",
		    '2': "ВТ",
		    '3': "СР",
		    '4': "ЧТ",
		    '5': "ПТ",
		    '6': "СБ",
		    '7': "ВС"}




def get_row(id):
		#
		# считывает строку с последним преподавателем
		#
		wb = load_workbook('journal.xlsx', read_only=True)
		sheet = wb['Teachers']
		# with open('last_row.txt', 'r') as file:
		# 	last_row = file.read()
		last_row = sheet.max_row
		cell_range = sheet["A2:A"+str(last_row)]
		for cellObj in cell_range:
			for cell in cellObj:				
				if str(cell.value) == str(id): # id имеет тип int, необходимо преобразование к строке
					return cell.row
def get_mark(name):
	print (name, who_is_absent)
	if name in set(who_is_absent):
		return "нет"
	else:
		return "x"

class Teacher():

	name = ''		
	id = ''
	subjects = []


	def __init__(self,id=id, name=name, subjects=subjects):
		self.name = ""#'Ne robit'
		self.id = id
		# self.subjects = []


		#
	

	def name_teacher(self, name=None, id=None):
		#
		# Устанавливает значения атрибутов в соответствие с данными в таблице "Teachers" в файле "journal.xlsx"
		#
	
		attributes = ['name']#, 'subjects']#Teacher.__dict__.keys()

		row = get_row(self.id)
		wb = load_workbook('journal.xlsx')
		col = 2
		sheet = wb['Teachers']
		for i in range(2, sheet.max_row + 1):
			if sheet.cell(column = 1, row = i).value == self.id:
				self.name = sheet.cell(column=2, row = i).value
				break

		# for attribute in attributes:
		# 	setattr(teacher, attribute, sheet.cell(column=col, row=row).value) # УСТАНАВЛИВАЕТ ЗНАЧЕНИЕ АТРИБУТОВ (объект, свойства класса, значение)
		# 	col+=1 # переход к следующему значению


class Lesson():
	teacher = ''
	subject = ''
	students = []


	def get_list(self, name_group, id):
		wb =load_workbook('journal.xlsx')
		sheet = wb[str(id)]
		col = 1
		row = 1
		students = []
		for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=1):
			for cell in row:
				if cell.value == name_group:
					cell_range = sheet[str(cell.column)+str(cell.row+1)+':'+str(cell.column)+"7"]
					for cellObj in cell_range:
						for cell_ in cellObj:
							if cell_.value!=None:
								students.append(cell_.value)
							else:
								break
					break
		return students

	# def keyboard_for_mark():	
	# 	keyboard = types.InlineKeyboardMarkup()
	# 	butns = []

	# 	for name in who_was:
	# 		butns.append(types.InlineKeyboardButton(text = name, callback_data = name))
	# 	butns.append(types.InlineKeyboardButton(text = 'Всё', callback_data = "Всё"))
	# 	butns.append(types.InlineKeyboardButton(text = 'Отмена', callback_data = 'Отмена'))
	# 	keyboard.add(*butns)
	# 	return keyboard

	def write_lesson(self, teacher=None, students=None):
		global current_subject
		global writing_date
		# print ("На запись")
		# global who_is_absent
		# print (type(writing_date))

		# writing_date = datetime.strptime(writing_date, "%d.%m.%y")
		lesson_list = []
		# answer=current_subject+", "+writing_date.strftime("%d.%m.%y")+"\n"
		# print (students)
		for name in students:
			lesson_list.append(writing_date)
			lesson_list.append(get_weekday(writing_date))
			lesson_list.append(get_month(writing_date))
			lesson_list.append(name)
			lesson_list.append(current_subject)
			lesson_list.append(get_mark(name))
			lesson_list.append(self.teacher)
			write_excel(lesson_list)
			lesson_list = []
		# if who_is_absent!=[]:
		# 	answer += "Отсутствовали:\n"
		# 	for i in who_is_absent:
		# 		answer+=i+';\n'
		# else:
		# 	answer += "Отсутствующих не было"
		# who_is_absent = []
		return "Запись прошла успешно"
	def __init__(self, teacher = teacher, subject = subject):
		self.teacher = teacher
		self.subject = subject



#### конец класса teacher
####

students = []

teacher = Teacher()

lesson = Lesson()

current_subject = ''

who_is_absent = []

who_was = []

sl_u_dates = {}

writing_date = ''

full_name_of_group = ''


lesson_object = {1: "", # название предмета 
				 2: [], # ученики
				 3: []} # дни недели

weeks_day = ["ПН", "ВТ", "СР", "ЧТ", "ПТ", "СБ", "ВС", "Подтвердить"]
timetable_obj = {}
weeks_day_copy = deepcopy(weeks_day)
name_group = ''
subj = []
struct_for_excel = {"Имя преподавателя": "",
					"Группы": []
					}
timetable_subjects = []
flag_for_handlers = 0
# flag == 1 - идёт запись по команде lesson
# flag == 2 - идёт запись по команде construct

step_for_construct = 0


@bot.message_handler(commands = ['construct'])
#setup_step_0 - инициализация команды construct, проверка id
def setup_step_0(text):
	global subjects
	global flag_for_handlers
	global step_for_construct
	global students
	students = []
	step_for_construct = 0
	try:
		bot.send_message(chat_id=text.chat.id, text = "Представьтесь",reply_markup = keyboard_subjects([]))
	except:
		step_for_construct+=1
		bot.edit_message_text(chat_id=text.message.chat.id,message_id = text.message.message_id, text = """Введите предмет и название группы через дефис. Например "Информатика-9 класс" """,
						 reply_markup = keyboard_subjects([]))
	step_for_construct+=1
	flag_for_handlers = 2





def preview():
	global current_subject
	global writing_date
	global who_is_absent
	writing_date = datetime.strptime(writing_date, "%d.%m.%y")
	answer=current_subject+", "+writing_date.strftime("%d.%m.%y")+"\n"
	if who_is_absent!=[]:
		answer += "Отсутствовали:\n"
		for i in who_is_absent:
			answer+=i+';\n'
	else:
		answer += "Отсутствующих не было"
	who_is_absent = []
	return answer

def get_date():
	return datetime.today()
def get_week_day():
	return weekday[str(datetime.date.isoweekday(datetime.today()))]
def get_weekday(i):
	return weekday[str(datetime.isoweekday(i))]
def get_month(i):

	return month_table[str(i.month)]



def write_excel(lesson_list):
		with open ("row.txt", "r") as count:
			row = count.read()
		row = int(row)
		wb = load_workbook('my_journal.xlsx')

		sheet = wb["Journal"]
		# print (lesson_list)
		for col in range(1, 8):			
			cell = sheet.cell(column =col, row=row, value=lesson_list[col-1])
			if col==1:
				cell.number_format = 'DD\.MM\.YY;@'
		row+=1
		wb.save('my_journal.xlsx')
		with open ("row.txt", 'w') as count:
				count.write(str(row))



@bot.message_handler(commands = ['start'])
def menu(text):
	#запуск бота, проверка по списку пользователей
	pass



@bot.message_handler(commands = ['file'])
def give_file(text):
	if text.from_user.id in (325726476 , 223103214):
		doc = open('my_journal.xlsx', 'rb')
		bot.send_document(chat_id= text.chat.id, data = doc)
		doc.close()

@bot.message_handler(commands = ['lesson'])
def msg(text):
	global teacher
	global sl_u_dates
	teacher.id = text.from_user.id
	teacher.name_teacher()
	lesson.teacher = teacher.name
	count = 1
	date = ''
	temp = []
	message = ''
	today = get_date()
	with open (str(text.from_user.id)+".txt", "r") as file:
		for line in file:
			date = line.split(" ")[0]
			subj = line[9:-1]
			if  datetime.strptime(date, '%d.%m.%y')<=today:
				message += str(count)+". "+ date+ " "+subj+"\n"
				temp.append(count)
				temp.append(date+ "-"+subj)
				sl_u_dates.update([temp])
				count+=1
				temp=[]
			else:
				break
	#bot.send_message(chat_id= text.chat.id, text = str_dates, reply_markup = keyboard_subjects(range(1, size+1)))
	if count !=1:
		global flag_for_handlers
		bot.send_message(chat_id=text.chat.id, text = message, reply_markup = keyboard_subjects(range(1, count)))
		flag_for_handlers = 1
	else:
		bot.send_message(chat_id=text.chat.id, text = "Нет доступных для записи занятий. Попробуйте позже.")


@bot.callback_query_handler(func=lambda text: True)
def func(text):
		global current_subject
		global who_was
		global who_is_absent
		global students
		global sl_u_dates
		global writing_date
		global flag_for_handlers
		global name_group
		if flag_for_handlers == 1:
			try:
				if int(text.data) in sl_u_dates:
					current_subject = sl_u_dates[int(text.data)].split("-")[1]
					writing_date = sl_u_dates[int(text.data)].split("-")[0]
					lesson.subject = current_subject
					name_group = current_subject +'-'+ sl_u_dates[int(text.data)].split("-")[2]
					students = lesson.get_list(name_group, teacher.id)
					who_was = deepcopy(students)
					bot.edit_message_text(chat_id = text.message.chat.id, message_id = text.message.message_id,
					 						text = 'На занятии были все?', reply_markup = keyboard_yes_no(2))
			except:

				if text.data == 'Отмена':

					bot.edit_message_text(chat_id = text.message.chat.id,
						 			message_id = text.message.message_id,
									text = "Отмена")
				elif text.data == 'Нет, не все':
					bot.edit_message_text(chat_id = text.message.chat.id,message_id = text.message.message_id,
										 text = "Кого не было?", reply_markup = keyboard_subjects(students))


				elif text.data == 'Да, все':		
					bot.edit_message_text(chat_id = text.message.chat.id, message_id = text.message.message_id,
												text = preview(), reply_markup = keyboard_subjects(["Подтвердить"]))

				elif text.data == "Всё":
					bot.edit_message_text(chat_id = text.message.chat.id, message_id = text.message.message_id,
												text = preview(), reply_markup = keyboard_subjects(["Подтвердить"]))


				elif text.data == "Занятия не было":
					who_was.remove(name)
					who_is_absent.append(name)
					bot.edit_message_text(chat_id = text.message.chat.id, message_id = text.message.message_id,
												text = preview(), reply_markup = keyboard_subjects(["Подтвердить"]))

				elif text.data == "Подтвердить":
					bot.edit_message_text(chat_id = text.message.chat.id, message_id = text.message.message_id,
										  text = lesson.write_lesson(teacher.name, students))
					f = open(str(text.from_user.id)+'.txt').read()
					f = f.replace(writing_date.strftime("%d.%m.%y")+" "+name_group+"\n",'')
					with open (str(text.from_user.id)+'.txt', "w") as file:
						file.write(f)

				else:
					if text.data in who_was:
						who_was.remove(text.data)
						who_is_absent.append(text.data)
						answer="Кого не было?\n"
						answer+=text.data
						bot.edit_message_text(chat_id = text.message.chat.id, message_id = text.message.message_id,
												 text = answer, reply_markup = keyboard_after_delete_name(who_was))
		elif flag_for_handlers == 2:
			global step_for_construct
			global struct_for_excel
			global weeks_day_copy
			global subj
			global timetable_obj
			if text.data == "Всё":
				bot.send_message(chat_id = text.from_user.id, text = "Введите фамилию и имя ученика",
								 reply_markup = keyboard_subjects(['Закончить ввод']))
				step_for_construct+=1
			elif text.data == "Закончить ввод":
				global group_info
				for i in students:
					group_info["Ученики"].append(i)
				bot.send_message(chat_id=text.from_user.id, text = "Выберите дни недели проведения этого занятия",
								 reply_markup = keyboard_subjects(weeks_day))
			elif text.data == "Подтвердить":
				struct_for_excel["Группы"].append(group_info)
				bot.edit_message_text(chat_id = text.message.chat.id, message_id = text.message.message_id,
									  text = "Выберите процедуру",
									  reply_markup = keyboard_subjects(["Добавить группу", "Составить расписание"]))

				# обнуление использованных переменных
				group_info = {"Название группы": "", 
			  				  "Ученики": [],
							  "Расписание":[]}
				weeks_day_copy = deepcopy(weeks_day)
				# внесение в базы данных
				with open (str(text.message.chat.id)+".json", "w") as file:
					dump(struct_for_excel, file, indent = 4)
				pass
			elif text.data == "Составить расписание":
				subj = []
				get_timetable(text.message.chat.id, timetable_obj)
				wb = load_workbook("journal.xlsx")
				sheet = wb['Teachers']
				count = 1
				for i in range(2, sheet.max_row):
					if sheet.cell(column=1, row=i).value == text.message.chat.id:

						sheet.cell(column = 1, row = i+1, value = text.message.chat.id)
						sheet.cell(column = 2, row = i+1, value = struct_for_excel["Имя преподавателя"])
						break
					else:
						count+=1
				if count == sheet.max_row:
					sheet.cell(column = 1, row = sheet.max_row+1, value = text.message.chat.id)
					sheet.cell(column = 2, row = sheet.max_row, value = struct_for_excel["Имя преподавателя"])
				try:
					wb.remove_sheet(wb[str(text.message.chat.id)])
					sheet = wb.create_sheet(str(text.message.chat.id))
				except:
					sheet = wb.create_sheet(str(text.message.chat.id))
				for j in range (1, len(struct_for_excel["Группы"])+1):
					sheet.cell(column = j, row = 1, value = struct_for_excel["Группы"][j-1]["Название группы"])
					for k in range(2, len(struct_for_excel["Группы"][j-1]["Ученики"])+2):
						sheet.cell(column = j, row = k, value = struct_for_excel["Группы"][j-1]["Ученики"][k-2])
				wb.save("journal.xlsx")
				bot.edit_message_text(chat_id=text.message.chat.id, message_id = text.message.message_id,
									  text = "Расписание составлено")
				# обнуление данных
				struct_for_excel = {"Имя преподавателя": "",
					  				"Группы": []}
				timetable_obj = {}
			elif text.data == "Добавить группу":
				students = []
				setup_step_0(text)
			elif text.data == "Отмена":
				bot.edit_message_text(chat_id=text.message.chat.id, message_id=text.message.message_id,
									  text = "Отмена")
				students = []
				step_for_construct=0
			elif text.data in weeks_day:
				group_info["Расписание"].append(text.data)
				weeks_day_copy.remove(text.data)
				subj = []
				try:
					subj.append(group_info["Название группы"])
					for i in timetable_obj[text.data]:
						subj.append(i)
				except:
					pass
				timetable_obj.update({text.data: subj})
				bot.edit_message_text(chat_id=text.message.chat.id, message_id=text.message.message_id,
									  text = "Выберите дни недели проведения этого занятия", reply_markup = keyboard_subjects(weeks_day_copy))
				

# struct_for_excel = {"Имя преподавателя": "", # записывается после представьтесь
# 					  "Группы": [
# 									
# 							]
# 					}
group_info = {"Название группы": "", 
			  "Ученики": [],
			  "Расписание":[]}

@bot.message_handler(content_types= 'text')
def create_timetable(text):
	global step_for_construct
	global flag_for_handlers
	global timetable_subjects
	global struct_for_excel
	global group_info
	if flag_for_handlers == 2:
		if step_for_construct == 1:
			struct_for_excel["Имя преподавателя"] = text.text
			bot.send_message(chat_id=text.chat.id, text = text.text+""", введите предмет и название группы через дефис. Например "Информатика-9 класс" """, reply_markup = keyboard_subjects([]))
			step_for_construct+=1
		elif step_for_construct==2:
			group_info["Название группы"] = text.text
			bot.send_message(chat_id=text.from_user.id, text = "Введите фамилию и имя ученика", reply_markup = keyboard_subjects(['Закончить ввод']))
			step_for_construct+=1
		elif step_for_construct == 3:
			students.append(text.text)
	pass
# bot.polling()

# bot.remove_webhook()

#  # Ставим заново вебхук
# bot.set_webhook(url=WEBHOOK_URL_BASE + WEBHOOK_URL_PATH,
#                 certificate=open(WEBHOOK_SSL_CERT, 'r'))
# cherrypy.config.update({
#     'server.socket_host': WEBHOOK_LISTEN,
#     'server.socket_port': WEBHOOK_PORT,
#     'server.ssl_module': 'builtin',
#     'server.ssl_certificate': WEBHOOK_SSL_CERT,
#     'server.ssl_private_key': WEBHOOK_SSL_PRIV
# })

#  # Собственно, запуск!
# cherrypy.quickstart(WebhookServer(), WEBHOOK_URL_PATH, {'/': {}})